"""
DocuAgent v3.1 — FastAPI Backend (Fixed)
=========================================
Fixes in this version:
  - Learning system: similarity-based feedback injection
  - Confidence: always returned, stored, displayed
  - /api/ai-insights endpoint added
  - Dashboard SQL interval bug fixed
  - /email-log auto-classifies incoming emails
"""
import os, json, uuid, logging, re
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import asyncpg, httpx, uvicorn
from fastapi import FastAPI, File, Form, UploadFile, HTTPException, Request
from fastapi.responses import HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
from pydantic import BaseModel

load_dotenv()

DB_URL         = os.getenv("DATABASE_URL", "postgresql://postgres:postgres@localhost:5432/docuagent")
QDRANT_URL     = os.getenv("QDRANT_URL",   "http://localhost:6333")
N8N_BASE_URL   = os.getenv("N8N_BASE_URL", "http://localhost:5678")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
COMPANY_NAME   = os.getenv("COMPANY_NAME",   "Agentify Kft.")
PORT           = int(os.getenv("PORT", "8000"))
CONF_THRESHOLD = float(os.getenv("CONFIDENCE_THRESHOLD", "0.70"))
_BASE          = Path(__file__).parent
UPLOAD_DIR     = _BASE / "uploads"
ALLOWED_EXTS   = {".pdf",".docx",".doc",".xlsx",".xls",".txt",".csv",".md"}

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-8s  %(message)s")
log = logging.getLogger("docuagent")

db_pool: Optional[asyncpg.Pool] = None

@asynccontextmanager
async def lifespan(app: FastAPI):
    global db_pool
    UPLOAD_DIR.mkdir(exist_ok=True)
    try:
        db_pool = await asyncpg.create_pool(DB_URL, min_size=2, max_size=10)
        log.info("PostgreSQL connected")
    except Exception as e:
        log.warning(f"PostgreSQL unavailable ({e})")
        db_pool = None
    yield
    if db_pool: await db_pool.close()

app = FastAPI(title="DocuAgent API", version="3.1", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

class ClassifyRequest(BaseModel):
    email_id: Optional[str] = None
    subject:  str
    body:     str
    sender:   Optional[str] = ""

class ClassifyResponse(BaseModel):
    can_answer:       bool
    confidence:       float
    category:         str
    reason:           str
    status:           str
    learned_override: bool = False

class ReplyRequest(BaseModel):
    email_id: Optional[str] = None
    subject: str; body: str; category: str

class FeedbackRequest(BaseModel):
    email_id: str; original_ai_decision: str; new_status: str; note: Optional[str] = ""

class StatusUpdateRequest(BaseModel):
    status: str; note: Optional[str] = ""

async def db_fetch(q, *a):
    if not db_pool: return []
    async with db_pool.acquire() as c: return await c.fetch(q, *a)

async def db_fetchrow(q, *a):
    if not db_pool: return None
    async with db_pool.acquire() as c: return await c.fetchrow(q, *a)

async def db_execute(q, *a):
    if not db_pool: return None
    async with db_pool.acquire() as c: return await c.execute(q, *a)

async def openai_chat(messages, max_tokens=800, json_mode=False):
    h = {"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"}
    b = {"model": "gpt-4o-mini", "messages": messages, "max_tokens": max_tokens}
    if json_mode: b["response_format"] = {"type": "json_object"}
    async with httpx.AsyncClient(timeout=30) as c:
        r = await c.post("https://api.openai.com/v1/chat/completions", headers=h, json=b)
        r.raise_for_status()
        return r.json()["choices"][0]["message"]["content"]

# ── Similarity ────────────────────────────────────────────────
def _normalize(t):
    t = t.lower()
    t = re.sub(r"[^\w\s]", " ", t)
    return re.sub(r"\s+", " ", t).strip()

def _similarity(a, b):
    if not a or not b: return 0.0
    stop = {"a","az","és","is","nem","de","the","is","and","to","of","in","for"}
    sa = set(_normalize(a).split()) - stop
    sb = set(_normalize(b).split()) - stop
    if not sa or not sb: return 0.0
    return len(sa & sb) / len(sa | sb)

# ── Learning ──────────────────────────────────────────────────
async def get_feedback_context(subject, body):
    """Returns (prompt_ctx, forced_status_or_None, similarity_score)"""
    rows = await db_fetch("""
        SELECT e.subject, e.body, e.category, f.user_decision, f.ai_decision
        FROM feedback f JOIN emails e ON e.id = f.email_id
        ORDER BY f.created_at DESC LIMIT 30
    """)

    best_match, best_score = None, 0.0
    combined = subject + " " + (body or "")[:500]

    for r in (rows or []):
        score = _similarity(combined, (r["subject"] or "") + " " + (r["body"] or "")[:500])
        if score > best_score:
            best_score = score
            best_match = r

    forced = None
    if best_match and best_score >= 0.60:
        forced = best_match["user_decision"]
        log.info(f"Learning override (sim={best_score:.2f}): {forced}")

    recent = await db_fetch("""
        SELECT f.ai_decision, f.user_decision, e.subject, e.category
        FROM feedback f JOIN emails e ON e.id = f.email_id
        ORDER BY f.created_at DESC LIMIT 10
    """)

    lines = []
    if recent:
        lines.append("\nPast human corrections (learn from these):")
        for r in recent:
            subj = r['subject'] or ''
            ai_d = r['ai_decision'] or ''
            ud   = r['user_decision'] or ''
            lines.append(f"  Subject: '{subj}'  AI: {ai_d} -> Human: {ud}")
    if best_match and best_score >= 0.35:
        ms = best_match['subject'] or ''
        mu = best_match['user_decision'] or ''
        lines.append(f"\nNOTE: Current email is {best_score:.0%} similar to '{ms}' -> corrected to {mu}. Weight this heavily.")

    return "\n".join(lines), forced, best_score


# ═══════════════════════════════ ENDPOINTS ═══════════════════

@app.post("/classify", response_model=ClassifyResponse)
async def classify_email(req: ClassifyRequest):
    if not OPENAI_API_KEY:
        return ClassifyResponse(can_answer=False, confidence=0.0, category="other",
                                reason="No API key", status="NEEDS_ATTENTION")

    feedback_ctx, forced, sim = await get_feedback_context(req.subject, req.body)

    if forced and sim >= 0.60:
        conf = round(0.50 + sim * 0.45, 2)
        can  = forced == "AI_ANSWERED"
        cat  = "complaint" if forced == "NEEDS_ATTENTION" else "inquiry"
        if req.email_id:
            await db_execute(
                "UPDATE emails SET category=$1,status=$2,ai_decision=$3,confidence=$4 WHERE id=$5",
                cat, forced,
                json.dumps({"can_answer":can,"confidence":conf,"reason":f"learned sim={sim:.2f}","learned_override":True}),
                conf, req.email_id)
        log.info(f"Classify LEARNED: {req.subject[:40]} → {forced} sim={sim:.2f}")
        return ClassifyResponse(can_answer=can, confidence=conf, category=cat,
                                reason=f"Learned ({sim:.0%} similar)", status=forced, learned_override=True)

    sys_prompt = f"""You are an email classifier for {COMPANY_NAME}.
Respond ONLY with valid JSON:
{{"can_answer":true/false,"confidence":0.0-1.0,"category":"complaint|inquiry|other","reason":"short reason"}}
Rules: confidence>={CONF_THRESHOLD} AND can_answer=true → auto-reply. Complaints → can_answer=false.{feedback_ctx}"""

    try:
        raw    = await openai_chat([{"role":"system","content":sys_prompt},
                                    {"role":"user","content":f"Subject: {req.subject}\n\n{req.body[:3000]}"}],
                                   max_tokens=300, json_mode=True)
        p      = json.loads(raw)
        can    = bool(p.get("can_answer", False))
        conf   = round(float(p.get("confidence", 0.0)), 2)
        cat    = p.get("category", "other")
        reason = p.get("reason", "")
        status = "AI_ANSWERED" if (can and conf >= CONF_THRESHOLD) else "NEEDS_ATTENTION"

        if req.email_id:
            await db_execute(
                "UPDATE emails SET category=$1,status=$2,ai_decision=$3,confidence=$4 WHERE id=$5",
                cat, status, json.dumps({"can_answer":can,"confidence":conf,"reason":reason}), conf, req.email_id)

        log.info(f"Classify GPT: {req.subject[:40]} → {status} conf={conf}")
        return ClassifyResponse(can_answer=can, confidence=conf, category=cat,
                                reason=reason, status=status)
    except Exception as e:
        log.error(f"Classify error: {e}")
        return ClassifyResponse(can_answer=False, confidence=0.0, category="other",
                                reason=str(e), status="NEEDS_ATTENTION")


@app.post("/generate-reply")
async def generate_reply(req: ReplyRequest):
    if not OPENAI_API_KEY: raise HTTPException(503, "No API key")
    sys = f"""Professional customer service for {COMPANY_NAME}.
Write helpful reply. Match language (Hungarian/English). No subject line."""
    try:
        reply = await openai_chat([{"role":"system","content":sys},
                                   {"role":"user","content":f"Category:{req.category}\nSubject:{req.subject}\n\n{req.body[:3000]}"}],
                                  max_tokens=500)
        if req.email_id:
            await db_execute("UPDATE emails SET ai_response=$1,status='AI_ANSWERED' WHERE id=$2", reply, req.email_id)
        return {"reply": reply, "email_id": req.email_id}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/feedback")
async def store_feedback(req: FeedbackRequest):
    valid = {"NEW","AI_ANSWERED","NEEDS_ATTENTION","CLOSED"}
    if req.new_status not in valid: raise HTTPException(400, f"Invalid: {valid}")
    await db_execute("UPDATE emails SET status=$1 WHERE id=$2", req.new_status, req.email_id)
    await db_execute("INSERT INTO feedback(email_id,ai_decision,user_decision,note) VALUES($1,$2,$3,$4)",
                     req.email_id, req.original_ai_decision, req.new_status, req.note or "")
    log.info(f"Feedback: {req.email_id} {req.original_ai_decision} → {req.new_status}")
    return {"status":"ok","email_id":req.email_id,"new_status":req.new_status}


@app.post("/email-log")
async def ingest_email(request: Request):
    data       = await request.json()
    email_id   = str(uuid.uuid4())
    message_id = data.get("message_id") or data.get("id") or email_id
    subject    = data.get("subject","")
    sender     = data.get("from", data.get("sender",""))
    body       = data.get("body", data.get("text",""))
    category   = data.get("category","other")
    urgent     = bool(data.get("urgent", False))
    ai_reply   = data.get("ai_reply","")

    existing = await db_fetchrow("SELECT id FROM emails WHERE message_id=$1", message_id)
    if existing:
        log.info(f"Duplicate skipped: {message_id}")
        return {"status":"duplicate","id":str(existing["id"])}

    try:
        await db_execute("""INSERT INTO emails(id,message_id,subject,sender,body,category,
                            status,urgent,ai_response,confidence,created_at)
                            VALUES($1,$2,$3,$4,$5,$6,'NEW',$7,$8,0.0,NOW())""",
                         email_id, message_id, subject, sender, body, category,
                         urgent, ai_reply or None)
    except Exception as e:
        log.error(f"Insert error: {e}")
        return {"status":"error","detail":str(e)}

    status = "NEW"; confidence = 0.0; learned = False
    if OPENAI_API_KEY:
        try:
            clf = await classify_email(ClassifyRequest(
                email_id=email_id, subject=subject, body=body or "", sender=sender))
            confidence = clf.confidence; status = clf.status; learned = clf.learned_override

            # Auto-generate AI reply if classified as AI_ANSWERED and no reply yet
            if status == "AI_ANSWERED" and not ai_reply:
                try:
                    reply_resp = await generate_reply(ReplyRequest(
                        email_id=email_id, subject=subject,
                        body=body or "", category=clf.category))
                    ai_reply = reply_resp.get("reply", "")
                    log.info(f"Auto-reply generated for: {subject[:50]}")
                except Exception as re:
                    log.warning(f"Auto-reply generation failed: {re}")
        except Exception as e:
            log.warning(f"Auto-classify failed: {e}")
            if ai_reply:
                status = "AI_ANSWERED"
                await db_execute("UPDATE emails SET status='AI_ANSWERED' WHERE id=$1", email_id)
    elif ai_reply:
        status = "AI_ANSWERED"
        await db_execute("UPDATE emails SET status='AI_ANSWERED' WHERE id=$1", email_id)

    log.info(f"Ingested: '{subject[:50]}' status={status} conf={confidence:.2f} learned={learned}")
    return {"status":"ok","id":email_id,"classified_status":status,
            "confidence":confidence,"learned_override":learned}


@app.get("/api/dashboard")
async def dashboard_data(days: int = 7):
    vectors = await _qdrant_count()
    n8n_ok  = await _check_n8n()

    rows = await db_fetch(f"""
        SELECT status, COUNT(*) AS cnt,
               COUNT(*) FILTER(WHERE urgent) AS urg,
               AVG(confidence) AS avg_conf
        FROM emails WHERE created_at > NOW() - INTERVAL '{days} days'
        GROUP BY status""")

    sm      = {r["status"]: r for r in (rows or [])}
    total   = sum(r["cnt"] for r in (rows or []))
    ai_ans  = sm.get("AI_ANSWERED",     {}).get("cnt", 0)
    needs   = sm.get("NEEDS_ATTENTION", {}).get("cnt", 0)
    closed  = sm.get("CLOSED",          {}).get("cnt", 0)
    new_cnt = sm.get("NEW",             {}).get("cnt", 0)
    urgent  = sum(r.get("urg", 0) for r in (rows or []))
    avg_c   = await db_fetchrow(f"SELECT AVG(confidence)*100 AS v FROM emails WHERE created_at>NOW()-INTERVAL '{days} days'")
    fb_cnt  = await db_fetchrow("SELECT COUNT(*) FROM feedback")

    timeline = await db_fetch(f"""
        SELECT DATE(created_at)::text AS day, COUNT(*) AS cnt,
               COUNT(*) FILTER(WHERE status='NEEDS_ATTENTION') AS needs
        FROM emails WHERE created_at>NOW()-INTERVAL '7 days'
        GROUP BY day ORDER BY day""")

    cats = await db_fetch(f"""
        SELECT COALESCE(category,'other') AS cat, COUNT(*) AS cnt
        FROM emails WHERE created_at>NOW()-INTERVAL '{days} days'
        GROUP BY cat""")
    cat_map = {r["cat"]: r["cnt"] for r in (cats or [])}

    act_rows = await db_fetch("""
        SELECT subject, sender, status, confidence, created_at
        FROM emails ORDER BY created_at DESC LIMIT 8""")
    activity = [{"type": "alert" if r["status"]=="NEEDS_ATTENTION" else "ok" if r["status"]=="AI_ANSWERED" else "email",
                 "title": (r["subject"] or "")[:60],
                 "meta":  f"{(r['sender'] or '')[:30]} · {r['status']} · {r['created_at'].strftime('%m-%d %H:%M') if r['created_at'] else ''}",
                 "confidence": round(float(r["confidence"] or 0)*100)}
                for r in (act_rows or [])]

    doc_rows = await db_fetch("SELECT id,filename,uploader,size_kb,lang,created_at,tag FROM documents ORDER BY created_at DESC LIMIT 10")
    docs = [{"id":str(r["id"]),"filename":r["filename"],"uploader":r["uploader"] or "—",
             "size_kb":r["size_kb"],"lang":r["lang"],
             "date":r["created_at"].strftime("%Y-%m-%d") if r["created_at"] else "",
             "ext":r["filename"].rsplit(".",1)[-1] if "." in (r["filename"] or "") else "?","tag":r["tag"]}
            for r in (doc_rows or [])]

    alerts = []
    needs_rate = round(needs / max(total,1)*100)
    urg_rate   = round(urgent / max(total,1)*100)
    if needs_rate >= 30: alerts.append({"type":"warn","message":f"<b>Figyelem:</b> NEEDS_ATTENTION arány {needs_rate}%."})
    if urg_rate  >= 40:  alerts.append({"type":"warn","message":f"<b>Urgent:</b> Sürgős emailek {urg_rate}%."})

    return {
        "meta": {"generated_at":datetime.now(timezone.utc).isoformat(),"range_days":days,
                 "n8n_status":"active" if n8n_ok else "offline","qdrant_vectors":vectors,
                 "openai_model":"gpt-4o-mini","company":COMPANY_NAME,"db_ok":db_pool is not None},
        "kpis": {"emails":{"value":total},"ai_answered":{"value":ai_ans},
                 "needs_attention":{"value":needs},"documents":{"value":len(docs)},
                 "avg_confidence":{"value":round(float(avg_c["v"] or 0)) if avg_c else 0},
                 "feedback_total":{"value":fb_cnt["count"] if fb_cnt else 0}},
        "status_breakdown":{"NEW":new_cnt,"AI_ANSWERED":ai_ans,"NEEDS_ATTENTION":needs,"CLOSED":closed},
        "charts":{"timeline":{"labels":[r["day"] for r in (timeline or [])],
                               "emails":[r["cnt"] for r in (timeline or [])],
                               "complaints":[r["needs"] for r in (timeline or [])]},
                  "category":{"complaint":cat_map.get("complaint",0),"inquiry":cat_map.get("inquiry",0),"other":cat_map.get("other",0)}},
        "activity":activity,"uploaders":[],"documents":docs,"alerts":alerts,
    }


@app.get("/api/emails")
async def list_emails(status: Optional[str]=None, limit: int=50, offset: int=0):
    if status:
        rows = await db_fetch("""SELECT id,subject,sender,body,category,status,urgent,confidence,ai_response,ai_decision,created_at
                                 FROM emails WHERE status=$1 ORDER BY created_at DESC LIMIT $2 OFFSET $3""",
                              status, limit, offset)
        total = await db_fetchrow("SELECT COUNT(*) FROM emails WHERE status=$1", status)
    else:
        rows = await db_fetch("""SELECT id,subject,sender,body,category,status,urgent,confidence,ai_response,ai_decision,created_at
                                 FROM emails ORDER BY created_at DESC LIMIT $1 OFFSET $2""", limit, offset)
        total = await db_fetchrow("SELECT COUNT(*) FROM emails")

    return {"emails":[{"id":str(r["id"]),"subject":r["subject"] or "","sender":r["sender"] or "",
                       "body":r["body"] or "","category":r["category"] or "other","status":r["status"],"urgent":r["urgent"],
                       "confidence":round(float(r["confidence"] or 0),2),"ai_response":r["ai_response"],
                       "ai_decision":r["ai_decision"],"created_at":r["created_at"].isoformat() if r["created_at"] else ""}
                      for r in (rows or [])],
            "total":total["count"] if total else 0, "limit":limit, "offset":offset}


@app.patch("/api/emails/{email_id}/status")
async def update_status(email_id: str, req: StatusUpdateRequest):
    valid = {"NEW","AI_ANSWERED","NEEDS_ATTENTION","CLOSED"}
    if req.status not in valid: raise HTTPException(400, f"Invalid: {valid}")
    row = await db_fetchrow("SELECT status,ai_decision FROM emails WHERE id=$1", email_id)
    if not row: raise HTTPException(404, "Not found")
    old = row["status"]
    ai  = json.dumps(row["ai_decision"]) if isinstance(row["ai_decision"],dict) else (row["ai_decision"] or old)
    await db_execute("UPDATE emails SET status=$1 WHERE id=$2", req.status, email_id)
    if old != req.status:
        await db_execute("INSERT INTO feedback(email_id,ai_decision,user_decision,note) VALUES($1,$2,$3,$4)",
                         email_id, ai, req.status, req.note or "")
        log.info(f"Learning: {email_id} {old} → {req.status}")
    return {"status":"ok","email_id":email_id,"new_status":req.status,"learning_stored":old!=req.status}


@app.get("/api/ai-insights")
async def ai_insights():
    if not OPENAI_API_KEY:
        return {"ai":{"problems":["OpenAI nem konfigurált"],"trends":["—"],"recommendations":["Állítsd be az OPENAI_API_KEY-t"]},
                "generated_at":datetime.now(timezone.utc).isoformat()}
    stats = await db_fetch("SELECT status,COUNT(*) AS c,AVG(confidence) AS ac FROM emails WHERE created_at>NOW()-INTERVAL '7 days' GROUP BY status")
    cats  = await db_fetch("SELECT COALESCE(category,'other') AS cat,COUNT(*) AS c FROM emails WHERE created_at>NOW()-INTERVAL '7 days' GROUP BY cat ORDER BY c DESC")
    fb    = await db_fetchrow("SELECT COUNT(*) FROM feedback")
    total = sum(r["c"] for r in (stats or []))
    sl = [f"  {r['status']}: {r['c']} db ({round(float(r['ac'] or 0)*100)}% conf)" for r in (stats or [])]
    cl = [f"  {r['cat']}: {r['c']}" for r in (cats or [])]
    prompt = f"""Analyze email data for {COMPANY_NAME}, return JSON insights in Hungarian.
Stats last 7 days: {chr(10).join(sl) or "No data"}
Categories: {chr(10).join(cl) or "No data"}
Total: {total}, Feedback corrections: {fb['count'] if fb else 0}
Return ONLY: {{"problems":["..."],"trends":["..."],"recommendations":["..."]}}"""
    try:
        raw  = await openai_chat([{"role":"user","content":prompt}], max_tokens=400, json_mode=True)
        data = json.loads(raw)
        return {"ai":data,"generated_at":datetime.now(timezone.utc).isoformat()}
    except Exception as e:
        return {"ai":{"problems":["Hiba: "+str(e)[:80]],"trends":[],"recommendations":[]},"generated_at":datetime.now(timezone.utc).isoformat()}


@app.post("/api/upload")
async def upload_doc(file: UploadFile=File(...), uploader_name: str=Form("Demo"),
                     uploader_email: str=Form("demo@agentify.hu"), tag: str=Form("general"),
                     department: str=Form("General"), access_level: str=Form("employee")):
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTS: raise HTTPException(400, f"Unsupported: {ext}")
    content = await file.read(); size_kb = round(len(content)/1024)
    dest = UPLOAD_DIR / f"{uuid.uuid4().hex[:8]}_{file.filename}"; dest.write_bytes(content)
    text = _extract_text(dest); lang = _detect_lang(text); doc_id = str(uuid.uuid4())
    qdrant_ok = False
    if OPENAI_API_KEY: qdrant_ok = await _store_in_qdrant(doc_id,file.filename,text,tag,department,access_level,uploader_email)
    await db_execute("INSERT INTO documents(id,filename,uploader,uploader_email,tag,department,access_level,size_kb,lang,qdrant_ok) VALUES($1,$2,$3,$4,$5,$6,$7,$8,$9,$10)",
                     doc_id,file.filename,uploader_name,uploader_email,tag,department,access_level,size_kb,lang,qdrant_ok)
    return {"status":"ok","id":doc_id,"filename":file.filename,"size_kb":size_kb,"lang":lang,"qdrant":qdrant_ok}


@app.get("/api/health")
async def health():
    qok,nok,v = False,False,-1
    try:
        async with httpx.AsyncClient(timeout=3) as c:
            r=await c.get(f"{QDRANT_URL}/healthz"); qok=r.status_code==200
            if qok: v=await _qdrant_count()
    except: pass
    try:
        async with httpx.AsyncClient(timeout=3) as c:
            r=await c.get(f"{N8N_BASE_URL}/healthz"); nok=r.status_code<500
    except: pass
    return {"status":"ok","db":{"ok":db_pool is not None},"qdrant":{"ok":qok,"vectors":v},
            "n8n":{"ok":nok},"openai":{"configured":bool(OPENAI_API_KEY)},"company":COMPANY_NAME,"version":"3.1"}


@app.get("/", response_class=HTMLResponse)
async def serve():
    for p in [_BASE/"dashboard.html", _BASE.parent/"dashboard"/"index.html"]:
        if p.exists(): return HTMLResponse(p.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>DocuAgent v3.1</h1><p><a href='/docs'>API Docs</a></p>")


async def _qdrant_count():
    try:
        async with httpx.AsyncClient(timeout=4) as c:
            r=await c.get(f"{QDRANT_URL}/collections/documents")
            return r.json().get("result",{}).get("points_count",0)
    except: return -1

async def _check_n8n():
    try:
        async with httpx.AsyncClient(timeout=3) as c:
            r=await c.get(f"{N8N_BASE_URL}/healthz"); return r.status_code<500
    except: return False

async def _store_in_qdrant(doc_id,filename,text,tag,dept,access,uploader):
    try:
        chunks=[text[i:i+1400] for i in range(0,min(len(text),12000),1400)]
        async with httpx.AsyncClient(timeout=60) as client:
            pts=[]
            for i,chunk in enumerate(chunks):
                r=await client.post("https://api.openai.com/v1/embeddings",
                    headers={"Authorization":f"Bearer {OPENAI_API_KEY}"},
                    json={"model":"text-embedding-3-small","input":chunk[:8000]})
                pts.append({"id":str(uuid.uuid4()),"vector":r.json()["data"][0]["embedding"],
                            "payload":{"filename":filename,"text":chunk,"tag":tag,"department":dept,
                                       "access_level":access,"uploader":uploader,"doc_id":doc_id,
                                       "chunk_index":i,"total_chunks":len(chunks),
                                       "upload_time":datetime.now(timezone.utc).isoformat()}})
            r2=await client.put(f"{QDRANT_URL}/collections/documents/points",json={"points":pts})
            return r2.status_code==200
    except Exception as e: log.warning(f"Qdrant failed: {e}"); return False

def _extract_text(path):
    ext=path.suffix.lower()
    try:
        if ext==".pdf":
            try:
                import fitz; doc=fitz.open(str(path)); t="\n".join(p.get_text() for p in doc); doc.close()
                if t.strip(): return t
            except ImportError: pass
            import pdfminer.high_level as pm; return pm.extract_text(str(path))
        elif ext in (".docx",".doc"):
            import docx; d=docx.Document(str(path)); return "\n".join(p.text for p in d.paragraphs if p.text.strip())
        elif ext in (".xlsx",".xls"):
            import openpyxl; wb=openpyxl.load_workbook(str(path),read_only=True,data_only=True)
            parts=[]
            for sn in wb.sheetnames:
                ws=wb[sn]
                for row in ws.iter_rows(values_only=True):
                    r=" | ".join(str(c) for c in row if c is not None)
                    if r.strip(): parts.append(r)
            wb.close(); return "\n".join(parts)
        else: return path.read_text(encoding="utf-8",errors="replace")
    except Exception as e: log.warning(f"Extract failed: {e}"); return ""

def _detect_lang(t):
    t=t.lower()
    hu=sum(1 for w in ["és","a","az","hogy","van","nem"] if f" {w} " in t)
    de=sum(1 for w in ["und","die","der","das","ist","nicht"] if f" {w} " in t)
    return "HU" if hu>de else ("DE" if de>0 else "EN")


if __name__=="__main__":
    uvicorn.run("main:app",host="0.0.0.0",port=PORT,reload=True)
